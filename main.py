import datetime
from typing import List
from fastapi import FastAPI
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def retorna_horarios(col,horario):
    
    return horario.idHora if col == 1 else horario.hora

# https://medium.com/data-hackers/como-manipular-planilhas-excel-com-o-python-6be8799f8dd7

class Campus(BaseModel):
    idCampus: str
    desc: str


class Local(BaseModel):
    idLocal: int
    desc: str
    idCampus: int

class Horarios(BaseModel):
    idHora: str
    hora: str
    ordem: int

class Data(BaseModel):
    campus: List[Campus]
    locais: List[Local]
    horarios: List[Horarios]

# comando para rodar : uvicorn main:app --reload
app = FastAPI()

@app.get("/")
async def root():
    return {"message": "Hello World"}

@app.post("/gerador_qts/")
async def gera_qts(data: Data):
    
    name_days_week = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira', 'Sexta-feira', 'Sábado', 'Domingo']
    # Obter a data atual
    today = datetime.date.today()

    # Obter o ano
    year = today.strftime('%Y')

    # Obter o dia da semana (0 = segunda-feira, 6 = domingo)
    weekday = today.weekday()

    # Calcular o primeiro dia da semana
    start_of_week = today - datetime.timedelta(days=weekday)

    # Calcular todos os dias da semana
    days_of_week = [start_of_week + datetime.timedelta(days=i) for i in range(7)]

    qtdCampus   = len(data.campus)
    qtdLocais   = len(data.locais)
    qtdHorarios = len(data.horarios)

    # Cria Novo workbook
    wb = Workbook()
    # Seleciona a active Sheet
    ws1 = wb.active
    # Rename it
    ws1.title = 'QTS_'+year
            
    # Colspan
    ws1['A1'] = 'Horário de início de aula'

    ws1.merge_cells('A1:B1')

    ws1['A2'] = 'Horário de início'
    ws1['B2'] = 'Tempo'


    # Imprimir todos os dias da semana
    countDays = 0
    for colDay in range(3,10):
        weekday = days_of_week[countDays].weekday()
        letterDay = get_column_letter(colDay)
        ws1[letterDay + '1'] = days_of_week[countDays].strftime('%d/%m')+' '+name_days_week[weekday]
        countDays += 1

    #Escreve alguns dados
    for col in range(1,3):
        for row in range(0,(qtdHorarios)):
            letter = get_column_letter(col)
            ws1[letter + str(row+3)] = retorna_horarios(col,data.horarios[row])

    # Salva arquivo (Se não colocar o caminho complete, ele salva
    # na mesma pasta do scritp.

    # Escreva os valores do array na planilha do Excel
   # for row in dados:
       # sheet.append(row)
    wb.save('QTS_'+year+'.xlsx')

    return data